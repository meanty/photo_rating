# -*- coding: utf-8 -*-
from django.core.serializers import base
from django.utils.encoding import smart_text
from openpyxl import Workbook as XLSX_Workbook # считывание и записывание в файлы.xlsx
from datetime import date, datetime, timedelta

class Serializer(base.Serializer):
    def start_serialization(self):
        """
        Start serialization -- open the xls document and the root element.
        """
        self.wb = XLSX_Workbook()
        self.count = 0

    def start_object(self, obj):
        if not hasattr(obj, "_meta"):
            raise base.SerializationError("Non-model object (%s) encountered during serialization" % type(obj))
        # count of row
        self.count += 1
        count_col = 1
        # flag saying that its first record in file
        if self.first:
            #creating sheet
            self.ws = self.wb.active
            self.ws.title = 'Clients dump'
            # names - dict that have fields of model,
            # his collumn in .xls file and verbose_name
            self.names = {}
            for field in obj._meta.fields: 
                if (self.options.has_key('name')) and (self.options['name'] == True):
                    #define verbose_name name
                    head_name = field.name
                else:
                    head_name = field.verbose_name
                    
                index = field.name
                self.names[index] = {}
                self.names[index]['verbose_name'] = head_name
                self.names[index]['column'] = count_col
                self.ws.cell(row = 1, column = count_col).value = head_name
                count_col += 1
        # write id value
        self.ws.cell(row = self.count+1, column = self.names['id']['column']).value = str(obj.id)    

    def handle_field(self, obj, field):
        valueString = ''
        if getattr(obj, field.name) is not None:
            if isinstance(field.value_to_string(obj), dict):
                valueString = field.value_to_string(obj)['human']
            else:
                valueString = field.value_to_string(obj)
        self.ws.cell(row = self.count+1, column = self.names[field.name]['column']).value = valueString


    def handle_fk_field(self, obj, field):
        pass

    def handle_m2m_field(self, obj, field):
        pass

    def end_serialization(self):
        self.wb.save(self.stream)

    def getvalue(self):
        return self.stream


def Deserializer(stream_or_string, **options):
    pass
